"""Product-Router: CRUD, Barcode-Lookup, Bulk-Import, Export."""
from __future__ import annotations

import csv
import io
import json
from datetime import date
from decimal import Decimal
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import Response, StreamingResponse
from sqlalchemy import func, or_, select
from sqlalchemy.orm import Session

from app.db.session import get_db
from app.models import Category, Product
from app.schemas import (
    ProductBulkRequest,
    ProductBulkResult,
    ProductCreate,
    ProductListResponse,
    ProductRead,
    ProductUpdate,
    decimal_to_cents,
)
from app.services.idempotency import (
    CachedResponse,
    client_op_id_header,
    get_cached_response,
    record_response,
)

router = APIRouter(prefix="/api/products", tags=["products"])


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _apply_payload(product: Product, payload: dict[str, Any]) -> None:
    """Überträgt ein gültiges Payload-Dict in das ORM-Objekt."""
    for key in ("sku", "barcode", "name", "category_id", "unit", "size_weight",
                "currency", "expiry_date", "supplier", "notes",
                "image_url", "color_tag", "is_active"):
        if key in payload and payload[key] is not None:
            setattr(product, key, payload[key])

    if "cost_price" in payload and payload["cost_price"] is not None:
        product.cost_price_cents = decimal_to_cents(payload["cost_price"])
    if "sell_price" in payload and payload["sell_price"] is not None:
        product.sell_price_cents = decimal_to_cents(payload["sell_price"])
    if "stock_quantity" in payload and payload["stock_quantity"] is not None:
        product.stock_quantity = Decimal(str(payload["stock_quantity"]))
    if "min_stock_level" in payload and payload["min_stock_level"] is not None:
        product.min_stock_level = Decimal(str(payload["min_stock_level"]))


def _validate_category(db: Session, category_id: int | None) -> None:
    if category_id is None:
        return
    if not db.get(Category, category_id):
        raise HTTPException(
            status_code=400,
            detail=f"Kategorie {category_id} existiert nicht.",
        )


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------


@router.get("", response_model=ProductListResponse)
def list_products(
    q: str | None = Query(None, description="Volltext-Suche über Name, SKU, Barcode"),
    category: int | None = Query(None, description="Kategorie-ID-Filter"),
    active: bool | None = Query(None, description="Aktiv-Filter (true/false)"),
    limit: int = Query(50, ge=1, le=500),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_db),
) -> ProductListResponse:
    stmt = select(Product)

    if q:
        like = f"%{q}%"
        stmt = stmt.where(
            or_(Product.name.ilike(like), Product.sku.ilike(like),
                Product.barcode.ilike(like))
        )
    if category is not None:
        stmt = stmt.where(Product.category_id == category)
    if active is not None:
        stmt = stmt.where(Product.is_active == active)

    # Total vor Pagination
    count_stmt = select(func.count()).select_from(stmt.subquery())
    total = db.execute(count_stmt).scalar_one()

    stmt = stmt.order_by(Product.name.asc()).limit(limit).offset(offset)
    rows = db.execute(stmt).scalars().all()

    return ProductListResponse(
        items=[ProductRead.from_orm_product(p) for p in rows],
        total=total,
        limit=limit,
        offset=offset,
    )


@router.get("/barcode/{code}", response_model=ProductRead)
def get_by_barcode(code: str, db: Session = Depends(get_db)) -> ProductRead:
    """Schneller Lookup per Barcode — genutzt von Scanner-UI."""
    stmt = select(Product).where(Product.barcode == code)
    product = db.execute(stmt).scalar_one_or_none()
    if product is None:
        raise HTTPException(status_code=404, detail=f"Kein Produkt mit Barcode '{code}'.")
    return ProductRead.from_orm_product(product)


@router.post("/bulk", response_model=ProductBulkResult, status_code=200)
def bulk_import(
    payload: ProductBulkRequest,
    db: Session = Depends(get_db),
) -> ProductBulkResult:
    """Bulk-Import aus JSON.

    Dedup-Strategie: vorhandenes Produkt wird per ``barcode`` (bevorzugt)
    oder ``sku`` gefunden und aktualisiert. Wenn weder barcode noch sku
    gesetzt sind, wird ein neues Produkt angelegt.
    """
    created = updated = skipped = 0
    errors: list[str] = []

    for idx, item in enumerate(payload.items):
        try:
            existing: Product | None = None
            if item.barcode:
                existing = db.execute(
                    select(Product).where(Product.barcode == item.barcode)
                ).scalar_one_or_none()
            if existing is None and item.sku:
                existing = db.execute(
                    select(Product).where(Product.sku == item.sku)
                ).scalar_one_or_none()

            data = item.model_dump()
            if existing is None:
                _validate_category(db, item.category_id)
                product = Product()
                _apply_payload(product, data)
                db.add(product)
                created += 1
            else:
                _validate_category(db, item.category_id)
                _apply_payload(existing, data)
                updated += 1
        except HTTPException as exc:
            errors.append(f"Item #{idx}: {exc.detail}")
            skipped += 1
        except Exception as exc:  # noqa: BLE001 — wir sammeln alle Fehler
            errors.append(f"Item #{idx}: {exc}")
            skipped += 1

    db.commit()
    return ProductBulkResult(
        created=created, updated=updated, skipped=skipped, errors=errors
    )


@router.post("/bulk/upload", response_model=ProductBulkResult)
async def bulk_upload(
    file: UploadFile = File(..., description="CSV-Datei mit Produkten"),
    db: Session = Depends(get_db),
) -> ProductBulkResult:
    """Bulk-Import aus einer CSV-Datei (Upload).

    Spaltennamen: ``barcode,sku,name,category,unit,size_weight,cost_price,
    sell_price,currency,stock_quantity,min_stock_level,expiry_date,
    supplier,notes,color_tag,is_active``.

    ``category`` kann Name oder ID sein. Preise als Dezimalzahl mit Punkt
    (z.B. ``1.99``).
    """
    raw = await file.read()
    text = raw.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))

    created = updated = skipped = 0
    errors: list[str] = []

    # Category-Name -> ID Cache
    cat_cache: dict[str, int] = {}

    def resolve_category_id(value: str | None) -> int | None:
        if not value:
            return None
        value = value.strip()
        if not value:
            return None
        if value.isdigit():
            return int(value)
        if value in cat_cache:
            return cat_cache[value]
        cat = db.execute(
            select(Category).where(Category.name == value)
        ).scalar_one_or_none()
        if cat is None:
            return None
        cat_cache[value] = cat.id
        return cat.id

    for idx, row in enumerate(reader):
        try:
            category_id = resolve_category_id(row.get("category"))

            item_data = {
                "barcode": (row.get("barcode") or None) or None,
                "sku": (row.get("sku") or None) or None,
                "name": row.get("name") or "",
                "category_id": category_id,
                "unit": row.get("unit") or "Stück",
                "size_weight": row.get("size_weight") or None,
                "cost_price": Decimal(row.get("cost_price") or "0"),
                "sell_price": Decimal(row.get("sell_price") or "0"),
                "currency": row.get("currency") or "EUR",
                "stock_quantity": Decimal(row.get("stock_quantity") or "0"),
                "min_stock_level": Decimal(row.get("min_stock_level") or "0"),
                "expiry_date": row.get("expiry_date") or None,
                "supplier": row.get("supplier") or None,
                "notes": row.get("notes") or None,
                "color_tag": row.get("color_tag") or "#3B82F6",
                "is_active": (row.get("is_active") or "true").lower()
                not in ("false", "0", "no", "nein"),
            }
            if not item_data["name"]:
                errors.append(f"Zeile {idx + 2}: leerer Name")
                skipped += 1
                continue

            existing = None
            if item_data["barcode"]:
                existing = db.execute(
                    select(Product).where(Product.barcode == item_data["barcode"])
                ).scalar_one_or_none()
            if existing is None and item_data["sku"]:
                existing = db.execute(
                    select(Product).where(Product.sku == item_data["sku"])
                ).scalar_one_or_none()

            if existing is None:
                _validate_category(db, category_id)
                product = Product()
                _apply_payload(product, item_data)
                db.add(product)
                created += 1
            else:
                _validate_category(db, category_id)
                _apply_payload(existing, item_data)
                updated += 1
        except Exception as exc:  # noqa: BLE001
            errors.append(f"Zeile {idx + 2}: {exc}")
            skipped += 1

    db.commit()
    return ProductBulkResult(
        created=created, updated=updated, skipped=skipped, errors=errors
    )


@router.get("/export", response_class=Response)
def export_products(
    format: str = Query("csv", pattern="^(csv|xlsx|pdf|json)$"),
    active: bool | None = Query(None),
    category: int | None = Query(None),
    db: Session = Depends(get_db),
):
    """Export der Preisliste in verschiedenen Formaten."""
    stmt = select(Product)
    if active is not None:
        stmt = stmt.where(Product.is_active == active)
    if category is not None:
        stmt = stmt.where(Product.category_id == category)
    stmt = stmt.order_by(Product.name.asc())
    rows: list[Product] = list(db.execute(stmt).scalars().all())

    if format == "json":
        data = [json.loads(ProductRead.from_orm_product(p).model_dump_json())
                for p in rows]
        return Response(
            content=json.dumps(data, indent=2, default=str),
            media_type="application/json",
            headers={
                "Content-Disposition": 'attachment; filename="products.json"'
            },
        )

    if format == "csv":
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow([
            "id", "sku", "barcode", "name", "category_id", "unit",
            "size_weight", "cost_price", "sell_price", "currency",
            "stock_quantity", "min_stock_level", "expiry_date",
            "supplier", "is_active", "color_tag",
        ])
        for p in rows:
            writer.writerow([
                p.id, p.sku or "", p.barcode or "", p.name,
                p.category_id or "", p.unit, p.size_weight or "",
                f"{p.cost_price_cents / 100:.2f}",
                f"{p.sell_price_cents / 100:.2f}",
                p.currency, p.stock_quantity, p.min_stock_level,
                p.expiry_date.isoformat() if p.expiry_date else "",
                p.supplier or "", "true" if p.is_active else "false",
                p.color_tag,
            ])
        return Response(
            content=buf.getvalue(),
            media_type="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": 'attachment; filename="products.csv"'
            },
        )

    if format == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Preisliste"
        headers = [
            "ID", "SKU", "Barcode", "Name", "Kategorie", "Einheit",
            "Größe", "EK (€)", "VK (€)", "Währung", "Bestand",
            "Min-Bestand", "MHD", "Lieferant", "Aktiv", "Farbe",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="3B82F6")
            cell.alignment = Alignment(horizontal="left")

        for p in rows:
            ws.append([
                p.id, p.sku, p.barcode, p.name, p.category_id, p.unit,
                p.size_weight,
                round(p.cost_price_cents / 100, 2),
                round(p.sell_price_cents / 100, 2),
                p.currency, float(p.stock_quantity),
                float(p.min_stock_level),
                p.expiry_date,
                p.supplier, p.is_active, p.color_tag,
            ])

        for col in ws.columns:
            length = max(
                (len(str(c.value)) for c in col if c.value is not None),
                default=10,
            )
            ws.column_dimensions[col[0].column_letter].width = min(
                length + 2, 40
            )

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return Response(
            content=out.read(),
            media_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            headers={
                "Content-Disposition": 'attachment; filename="products.xlsx"'
            },
        )

    # PDF
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4, leftMargin=15 * mm, rightMargin=15 * mm
    )
    styles = getSampleStyleSheet()
    elems: list = []
    today = date.today().isoformat()
    elems.append(Paragraph("MarktPilot — Preisliste", styles["Title"]))
    elems.append(Paragraph(f"Stand: {today}", styles["Normal"]))
    elems.append(Spacer(1, 6 * mm))

    table_data = [["Name", "Barcode", "VK", "Bestand", "Einheit"]]
    for p in rows:
        # Deutsches Format: Komma als Dezimaltrenner, " €" mit Leerzeichen.
        vk_euro = f"{p.sell_price_cents / 100:.2f}".replace(".", ",") + " €"
        table_data.append([
            p.name,
            p.barcode or "",
            vk_euro,
            str(p.stock_quantity),
            p.unit,
        ])
    table = Table(table_data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3B82F6")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (2, 1), (2, -1), "RIGHT"),
                ("ALIGN", (3, 1), (3, -1), "RIGHT"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                 [colors.whitesmoke, colors.HexColor("#EFF6FF")]),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ]
        )
    )
    elems.append(table)
    doc.build(elems)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type="application/pdf",
        headers={
            "Content-Disposition": 'attachment; filename="products.pdf"'
        },
    )


@router.get("/{product_id}", response_model=ProductRead)
def get_product(product_id: int, db: Session = Depends(get_db)) -> ProductRead:
    product = db.get(Product, product_id)
    if product is None:
        raise HTTPException(status_code=404, detail="Produkt nicht gefunden.")
    return ProductRead.from_orm_product(product)


@router.post("", response_model=ProductRead, status_code=201)
def create_product(
    payload: ProductCreate,
    db: Session = Depends(get_db),
    client_op_id: str | None = Depends(client_op_id_header),
) -> ProductRead:
    # --- Idempotenz: gleicher X-Client-Op-Id → gespeicherte Antwort zurück ---
    if client_op_id:
        cached = get_cached_response(db, client_op_id, "POST /api/products")
        if cached is not None:
            replay = CachedResponse(cached.response_json, cached.status_code)
            replay.raise_for_status()
            return ProductRead.model_validate(replay.body())
    # --- Ende Idempotenz-Vorprüfung ---

    _validate_category(db, payload.category_id)

    # Eindeutigkeit sicherstellen
    if payload.barcode:
        exists = db.execute(
            select(Product).where(Product.barcode == payload.barcode)
        ).scalar_one_or_none()
        if exists:
            raise HTTPException(
                status_code=409,
                detail=f"Barcode '{payload.barcode}' ist bereits vergeben.",
            )
    if payload.sku:
        exists = db.execute(
            select(Product).where(Product.sku == payload.sku)
        ).scalar_one_or_none()
        if exists:
            raise HTTPException(
                status_code=409,
                detail=f"SKU '{payload.sku}' ist bereits vergeben.",
            )

    product = Product()
    _apply_payload(product, payload.model_dump())
    db.add(product)
    db.commit()
    db.refresh(product)
    result = ProductRead.from_orm_product(product)

    # Idempotenz-Antwort cachen
    if client_op_id:
        record_response(
            db,
            client_op_id=client_op_id,
            endpoint="POST /api/products",
            status_code=201,
            response_body=result.model_dump(mode="json"),
        )
    return result


@router.put("/{product_id}", response_model=ProductRead)
def update_product(
    product_id: int,
    payload: ProductUpdate,
    db: Session = Depends(get_db),
) -> ProductRead:
    product = db.get(Product, product_id)
    if product is None:
        raise HTTPException(status_code=404, detail="Produkt nicht gefunden.")

    if payload.category_id is not None:
        _validate_category(db, payload.category_id)

    # Barcode-Konflikt mit anderem Produkt?
    if payload.barcode and payload.barcode != product.barcode:
        exists = db.execute(
            select(Product).where(
                Product.barcode == payload.barcode, Product.id != product_id
            )
        ).scalar_one_or_none()
        if exists:
            raise HTTPException(
                status_code=409,
                detail=f"Barcode '{payload.barcode}' ist bereits vergeben.",
            )

    if payload.sku and payload.sku != product.sku:
        exists = db.execute(
            select(Product).where(
                Product.sku == payload.sku, Product.id != product_id
            )
        ).scalar_one_or_none()
        if exists:
            raise HTTPException(
                status_code=409, detail=f"SKU '{payload.sku}' ist bereits vergeben."
            )

    _apply_payload(product, payload.model_dump(exclude_unset=True))
    db.commit()
    db.refresh(product)
    return ProductRead.from_orm_product(product)


@router.delete("/{product_id}", response_model=ProductRead)
def delete_product(product_id: int, db: Session = Depends(get_db)) -> ProductRead:
    """Soft-Delete: setzt nur is_active=false, löscht keine Daten."""
    product = db.get(Product, product_id)
    if product is None:
        raise HTTPException(status_code=404, detail="Produkt nicht gefunden.")
    product.is_active = False
    db.commit()
    db.refresh(product)
    return ProductRead.from_orm_product(product)